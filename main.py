import io
from fastapi import FastAPI, Depends, File, UploadFile, Form, Request, HTTPException, Path, Query
from fastapi.responses import RedirectResponse, Response, FileResponse, JSONResponse, HTMLResponse, StreamingResponse
from starlette.middleware.sessions import SessionMiddleware
import os
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from database import SessionLocal, Issue, IssueDetail, Base, ReviewProposal, LawUpdate, engine
from datetime import date
from sqlalchemy import Column, Integer, String, Date, func, extract, desc, and_, or_, text
import crud
from pydantic import BaseModel
import shutil
import mimetypes
from urllib.parse import quote, unquote
from fastapi.staticfiles import StaticFiles
from typing import List, Optional
from jinja2 import Environment
import json
import pandas as pd
from fastapi.responses import FileResponse
from datetime import datetime
import csv
from database import get_db, TestReview, Physicochemical, Efficacy, Residue, Toxicity, Supplement
from starlette.responses import RedirectResponse
from fastapi.exception_handlers import http_exception_handler
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.responses import PlainTextResponse
from datetime import timedelta
from openpyxl import Workbook
from io import BytesIO
from sqlalchemy import text
import openpyxl
import shutil
from fastapi import APIRouter
from typing import Optional
from database import Schedule, ScheduleHistory
from datetime import datetime
from database import ActionLog, LoginLog
from fastapi.responses import StreamingResponse
from database import Base, engine
from database import IssueAlertSubscription
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
from datetime import datetime
import pytz
from sqlalchemy import func
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
Base.metadata.create_all(bind=engine)



def get_updated_issues_today(db: Session):
    # 한국 시간 기준으로 오늘 날짜만 추출
    korea_today = datetime.now(pytz.timezone("Asia/Seoul")).date()

    # 오늘 updated_at이 변경된 이슈
    updated_issues = (
        db.query(Issue)
        .filter(func.date(Issue.updated_at) == korea_today)
        .all()
    )

    # 오늘 등록된 진행사항이 있는 이슈 ID 목록
    recent_details = (
        db.query(IssueDetail.issue_id)
        .filter(func.date(IssueDetail.date) == korea_today)
        .distinct()
        .all()
    )
    issue_ids = [r.issue_id for r in recent_details]

    # 해당 ID에 해당하는 이슈들 조회
    issues_from_details = db.query(Issue).filter(Issue.id.in_(issue_ids)).all()

    # 두 결과를 합쳐서 중복 제거 후 반환
    combined = {issue.id: issue for issue in (updated_issues + issues_from_details)}
    return list(combined.values())



def send_email(to_email: str, subject: str, html_content: str):
    import smtplib
    from email.mime.text import MIMEText
    from email.utils import formataddr

    smtp_server = "smtps.hiworks.com"     # ✅ 정확히 이 주소
    smtp_port = 465                       # ✅ SSL 포트
    sender_email = "choss@koreacpa.org"
    sender_name = "농약 이슈 시스템"
    password = "1adYlptkAGReyKM1qU0t"  # ✅ 반드시 메일 전용 비번 사용!

    msg = MIMEText(html_content, "html")
    msg["Subject"] = subject
    msg["From"] = formataddr((sender_name, sender_email))
    msg["To"] = to_email

    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, to_email, msg.as_string())
            print("✅ 메일 발송 성공")
    except Exception as e:
        print("❌ 메일 발송 실패:", e)
        raise

def send_daily_update_emails(db: Session):
    updated_issues = get_updated_issues_today(db)

    if not updated_issues:
        return  # 오늘 업데이트된 이슈가 없으면 발송 생략

    # 메일 본문 생성
    html = "<h3>오늘 업데이트된 이슈 목록</h3><ul>"
    for issue in updated_issues:
        html += f"<li>{issue.issue_name} ({issue.updated_at})</li>"
    html += "</ul>"

    # 알림 구독자 가져오기
    subscribers = db.query(IssueAlertSubscription).all()
    for sub in subscribers:
        send_email(
            to_email=sub.email,
            subject="📢 농약 이슈 업데이트 알림",
            html_content=html
        )

def add_columns_if_not_exist():
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE test_reviews ADD COLUMN institution_name TEXT"))
        except:
            pass
        try:
            conn.execute(text("ALTER TABLE test_reviews ADD COLUMN review_type TEXT"))
        except:
            pass

add_columns_if_not_exist()

class ScheduleCreate(BaseModel):
    start_date: date
    end_date: Optional[date] = None
    title: str
    location: Optional[str] = None


class SupplementResponseUpdate(BaseModel):
    response_method: str
    response_result: str

# ✅ 관리자 여부 확인 함수 추가
def is_admin(request: Request):
    user = request.session.get("user")
    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return user

def get_current_user_required(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    return user

def log_action(db, username: str, action_type: str, table: str, target_id: int):
    log = ActionLog(
        username=username,
        action_type=action_type,
        target_table=table,
        target_id=target_id,
        timestamp=datetime.utcnow()
    )
    db.add(log)
    db.commit()

def log_login(db, username: str, ip: str, success: bool):
    log = LoginLog(
        username=username,
        ip_address=ip,
        success=success,
        timestamp=datetime.utcnow()
    )
    db.add(log)
    db.commit()

# ✅ FastAPI 앱 초기화
app = FastAPI()

@app.get("/admin/export-logins")
def export_login_logs(db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    logs = db.query(LoginLog).order_by(LoginLog.timestamp.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["timestamp", "username", "ip_address", "success"])

    for log in logs:
        writer.writerow([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.username,
            log.ip_address,
            "Success" if log.success else "Fail"
        ])

    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={
        "Content-Disposition": "attachment; filename=login_logs.csv"
    })

@app.delete("/admin/logins/cleanup")
def delete_old_logins(db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    threshold = datetime.utcnow() - timedelta(days=30)
    deleted = db.query(LoginLog).filter(LoginLog.timestamp < threshold).delete()
    db.commit()
    return {"deleted": deleted}

@app.get("/favicon.ico")
async def favicon():
    return FileResponse("static/favicon.ico")

# ✅ 데이터베이스 세션 의존성 설정
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ✅ `SessionMiddleware`를 가장 먼저 추가 (미들웨어 순서 강제 적용)
app.add_middleware(
    SessionMiddleware,
    secret_key="super_secure_key_12345",
    session_cookie="session_id",
    max_age=1800,
    same_site="lax",
    https_only=False,
)

# ✅ Jinja2 템플릿 설정
templates = Jinja2Templates(directory="templates")

templates.env.filters["loads"] = json.loads

@app.exception_handler(StarletteHTTPException)
async def custom_http_exception_handler(request: Request, exc: StarletteHTTPException):
    # 🔹 요청 경로가 HTML 페이지라면
    if exc.status_code == 401 and request.headers.get("accept", "").find("text/html") >= 0:
        return templates.TemplateResponse("not_logged_in.html", {"request": request}, status_code=401)

    # 🔹 API나 기타 요청은 JSON 또는 기본 응답
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

# ✅ 하드코딩된 사용자 계정
USERS = {
    "adminkcpa": {"password": "admin123", "role": "admin", "display_name": "관리자"},

    # 회사별 사용자 (11개)
    "KyungNong": {"password": "user123", "role": "user", "display_name": "경농"},
    "NongHyupChemical": {"password": "user123", "role": "user", "display_name": "농협케미컬"},
    "DongBangAgro": {"password": "user123", "role": "user", "display_name": "동방아그로"},
    "BayerCropScience": {"password": "user123", "role": "user", "display_name": "바이엘크롭사이언스"},
    "SungBoChemical": {"password": "user123", "role": "user", "display_name": "성보화학"},
    "SyngentaKorea": {"password": "user123", "role": "user", "display_name": "신젠타코리아"},
    "SunmoonGreenScience": {"password": "user123", "role": "user", "display_name": "선문그린사이언스"},
    "Enbio": {"password": "user123", "role": "user", "display_name": "인바이오"},
    "FarmHannong": {"password": "user123", "role": "user", "display_name": "팜한농"},
    "HanKookSamgong": {"password": "user123", "role": "user", "display_name": "한국삼공"},
    "HanearlScience": {"password": "user123", "role": "user", "display_name": "한얼싸이언스"},
}

# ✅ 로그인 여부 확인 함수 (이제 `Depends()` 방식으로 처리)
def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        return None
    return user

# ✅ 로그인 후 이동할 선택 페이지
@app.get("/select-menu", response_class=HTMLResponse)
async def select_menu(request: Request, user: dict = Depends(get_current_user_required)):
    expires_str = user.get("expires_at")
    if expires_str:
        expires_at = datetime.fromisoformat(expires_str)
        remaining_seconds = max(int((expires_at - datetime.utcnow()).total_seconds()), 0)
    else:
        remaining_seconds = 0  # 만료시간이 없으면 기본값

    return templates.TemplateResponse("select_menu.html", {
        "request": request,
        "session_seconds": remaining_seconds,
        "username": user["username"],
        "user": user
    })

# ✅ 농약 현안 관리 → 기존 이슈 목록 페이지
@app.get("/issues/", response_class=HTMLResponse)
async def issue_list_page(request: Request, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    is_admin = user["role"] == "admin"
    username = user["username"]

    if is_admin:
        issues = db.query(Issue).all()
    else:
        issues = db.query(Issue).filter(
            or_(
                Issue.is_hidden == False,
                and_(
                    Issue.is_hidden == True,
                    Issue.authorized_users != None,
                    Issue.authorized_users != "",
                    Issue.authorized_users.contains(f",{username},")
                )
            )
        ).all()

    # 🔹 최근 5일 이내 업데이트 여부 표시
    for issue in issues:
        issue.is_recent_update = (date.today() - issue.updated_at).days <= 5

    return templates.TemplateResponse("index.html", {
        "request": request,
        "issues": issues,
        "username": user["username"],
        "user": user
    })

@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    else:
        return RedirectResponse(url="/select-menu", status_code=302)

# 로그인 페이지 렌더링
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    # 'admin' 계정은 드롭다운에서 제외
    company_users = {
        key: value for key, value in USERS.items()
        if value["role"] == "user"
    }
    return templates.TemplateResponse("login.html", {
        "request": request,
        "company_users": company_users
    })

# 로그인 처리
@app.post("/login")
def login(request: Request, db: Session = Depends(get_db), username: str = Form(...), password: str = Form(...)):
    user = USERS.get(username)

    if not user or user["password"] != password:
        log_login(db, username=username, ip=request.client.host, success=False)
        company_users = {
            k: v for k, v in USERS.items()
            if k != "admin" and v.get("display_name") != "준회원(공통)"
        }
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "아이디 또는 비밀번호가 잘못되었습니다.",
            "company_users": company_users
        })

    print("✅ 로그인 성공:", username)  # 확인용
    log_login(db, username=username, ip=request.client.host, success=True)

    expires_at = datetime.utcnow() + timedelta(seconds=1800)
    request.session["user"] = {
        "username": username,
        "role": user["role"],
        "display_name": user.get("display_name", username),
        "expires_at": expires_at.isoformat()
    }

    return RedirectResponse(url="/select-menu", status_code=302)

@app.post("/admin-login")
def admin_login(request: Request, db: Session = Depends(get_db), username: str = Form(...), password: str = Form(...)):
    user = USERS.get(username)
    if not user or user["role"] != "admin" or user["password"] != password:
        log_login(db, username=username, ip=request.client.host, success=False)
        return templates.TemplateResponse("admin_login.html", {
            "request": request,
            "error": "관리자 아이디 또는 비밀번호가 잘못되었습니다."
        })

    # ✅ 로그인 성공 기록
    log_login(db, username=username, ip=request.client.host, success=True)

    request.session["user"] = {
        "username": username,
        "role": user["role"],
        "display_name": user.get("display_name", username),
        "expires_at": (datetime.utcnow() + timedelta(seconds=1800)).isoformat()
    }
    return RedirectResponse(url="/select-menu", status_code=302)
    
@app.get("/current-user")
def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    return {
    "username": user["username"],
    "role": user["role"],
    "expires_at": user.get("expires_at")  # 🔥 이거 꼭 포함!
}

# 로그아웃 처리
@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session_id")  # FastAPI 기본 세션
    response.delete_cookie("session")     # 🔥 브라우저에 남아 있는 session 쿠키도 삭제
    return response

# 📌 업로드된 파일 저장 경로
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 📌 요청 데이터 검증을 위한 Pydantic 모델
class IssueCreate(BaseModel):
    issue_number: str
    issue_name: str
    issue_date: date
    response_period: str
    category: str
    response_team: str
    government_officials: str
    business_impact: str
    kpi: str
    issue_end_date: Optional[str]  # 🔥 문자열 또는 미정 허용
    stakeholders: str
    result_summary: str
    completion_status: str
    other_remarks: str

# 📌 이슈 추가 API
@app.post("/issues/")
def create_issue(issue: IssueCreate, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    # 🔥 이슈 종료일 처리 (미정인 경우 None으로 저장)
    if issue.issue_end_date and issue.issue_end_date.strip() != "미정":
        try:
            parsed_date = datetime.strptime(issue.issue_end_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="이슈 종료일 형식이 잘못되었습니다. (예: 2025-12-31 또는 미정)")
    else:
        parsed_date = None

    # ✅ 나머지 필드는 그대로 전달, issue_end_date만 별도 처리
    new_issue = Issue(
        **{**issue.dict(exclude={"issue_end_date"}), "issue_end_date": parsed_date},
        updated_at=date.today()
    )

    db.add(new_issue)
    db.commit()
    db.refresh(new_issue)
    return {"message": "이슈가 성공적으로 추가되었습니다.", "data": new_issue}

# 📌 모든 이슈 조회 API
@app.get("/api/issues")
def read_issues(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    username = user["username"]
    is_admin = user and user.get("role") == "admin"

    if is_admin:
        issues = db.query(Issue).all()
    else:
        issues = db.query(Issue).filter(
            or_(
                Issue.is_hidden == False,
                and_(
                    Issue.is_hidden == True,
                    Issue.authorized_users != None,
                    Issue.authorized_users != "",
                    Issue.authorized_users.contains(f",{username},")
                )
            )
        ).all()

    for issue in issues:
        issue.is_recent_update = (date.today() - issue.updated_at).days <= 5

    return {"message": "이슈 목록", "data": issues}

@app.get("/issues-data")
def get_issues_data(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    is_admin = user and user.get("role") == "admin"
    username = user["username"]

    if is_admin:
        issues = db.query(Issue).all()
    else:
        issues = db.query(Issue).filter(
            or_(
                Issue.is_hidden == False,
                and_(
                    Issue.is_hidden == True,
                    Issue.authorized_users != None,
                    Issue.authorized_users != "",
                    Issue.authorized_users.like(f"%,{username},%")
                )
            )
        ).all()

    issue_list = []
    for issue in issues:
        issue_list.append({
            "id": issue.id,
            "issue_number": issue.issue_number,
            "issue_name": issue.issue_name,
            "issue_date": issue.issue_date.isoformat() if issue.issue_date else "",
            "response_period": issue.response_period,
            "category": issue.category,
            "response_team": issue.response_team,
            "government_officials": issue.government_officials,
            "business_impact": issue.business_impact,
            "kpi": issue.kpi,
            "issue_end_date": issue.issue_end_date.isoformat() if issue.issue_end_date else "",
            "stakeholders": issue.stakeholders,
            "completion_status": issue.completion_status,
            "updated_at": issue.updated_at.isoformat() if issue.updated_at else "",
            "is_hidden": issue.is_hidden,
        })

    return {"data": issue_list}

# 📌 특정 이슈 조회 API
@app.get("/issues/{issue_id}")
def read_issue(issue_id: int, db: Session = Depends(get_db)):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        return {"message": "이슈를 찾을 수 없습니다."}
    return {"message": "이슈 상세 정보", "data": issue}

# 📌 특정 이슈 삭제 API
@app.delete("/issues/{issue_id}")
def delete_issue(issue_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="이슈를 찾을 수 없습니다.")

    db.delete(issue)
    db.commit()
    return {"message": "이슈가 성공적으로 삭제되었습니다."}

# 📌 세부 진행 사항 추가 API (파일 업로드 포함)
@app.post("/issues/{issue_id}/details")
def add_issue_detail(
    issue_id: int,
    date: date = Form(...),
    content: str = Form(...),
    files: List[UploadFile] = File(None),  # ✅ 여러 개 파일 받기
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="이슈를 찾을 수 없습니다.")

    # 🔥 이슈명으로 폴더 경로 설정
    issue_folder = os.path.join(UPLOAD_DIR, issue.issue_name)
    os.makedirs(issue_folder, exist_ok=True)  # 폴더가 없으면 생성

    file_paths = []
    if files:
        for file in files:
            file_path = os.path.join(issue_folder, file.filename)  # ✅ 폴더 경로 포함
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            file_paths.append(f"{issue.issue_name}/{file.filename}")  # ✅ 경로를 상대 경로로 저장

    new_detail = IssueDetail(
        issue_id=issue_id,
        date=date,
        content=content,
        file_path=json.dumps(file_paths)  # ✅ 여러 개의 파일을 JSON 리스트로 저장
    )
    db.add(new_detail)
    db.commit()
    db.refresh(new_detail)

    return {"message": "세부 진행 사항이 추가되었습니다.", "data": new_detail}

# 📌 세부 진행 사항 목록 조회 (날짜순 정렬)
@app.get("/issues/{issue_id}/details")
def get_issue_details(issue_id: int, db: Session = Depends(get_db)):
    details = db.query(IssueDetail).filter(IssueDetail.issue_id == issue_id).order_by(IssueDetail.date.asc()).all()
    
    result = []
    for d in details:
        file_paths = json.loads(d.file_path) if d.file_path else []  # ✅ 파일 리스트 복원
        result.append({
            "id": d.id,
            "date": d.date,
            "content": d.content,
            "file_paths": file_paths
        })
    
    return {"message": "세부 진행 사항 목록", "data": result}
# 📌 세부 진행 사항 삭제 API
@app.delete("/issues/detail/{detail_id}")
def delete_issue_detail(detail_id: int, db: Session = Depends(get_db)):
    detail = db.query(IssueDetail).filter(IssueDetail.id == detail_id).first()
    if not detail:
        return {"message": "세부 진행 사항을 찾을 수 없습니다."}

    db.delete(detail)
    db.commit()
    return {"message": "세부 진행 사항이 삭제되었습니다."}

# 📌 세부 진행 사항 수정 API
class IssueDetailUpdate(BaseModel):
    date: str
    content: str

@app.put("/issues/detail/{detail_id}")
def update_issue_detail(detail_id: int, detail_update: IssueDetailUpdate, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    detail = db.query(IssueDetail).filter(IssueDetail.id == detail_id).first()
    if not detail:
        raise HTTPException(status_code=404, detail="세부 진행 사항을 찾을 수 없습니다.")

    # 🔧 문자열을 date 객체로 변환
    detail.date = datetime.strptime(detail_update.date, "%Y-%m-%d").date()
    detail.content = detail_update.content

    db.commit()
    db.refresh(detail)
    return {"message": "세부 진행 사항이 수정되었습니다.", "data": detail}

# 📌 파일 다운로드 API (MIME 타입 자동 감지)
@app.get("/files/{filename:path}")
def download_file(filename: str):
    # ✅ 한글 파일명 처리 (URL 디코딩)
    decoded_filename = unquote(filename, encoding='utf-8')

    # ✅ 파일 경로 설정
    file_path = os.path.join(UPLOAD_DIR, decoded_filename)

    # ✅ 파일 존재 여부 확인
    if not os.path.isfile(file_path):
        return Response(content=f"파일을 찾을 수 없습니다: {decoded_filename}", status_code=404)

    # ✅ MIME 타입 자동 감지
    media_type, _ = mimetypes.guess_type(file_path)
    if media_type is None:
        media_type = "application/octet-stream"

    # ✅ 한글 파일명을 안전하게 브라우저에서 다운로드하도록 인코딩 처리
    encoded_filename = quote(decoded_filename, encoding='utf-8')

    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=decoded_filename,  # 🔥 파일명이 깨지지 않도록 설정
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

# 📌 정적 파일 서빙
app.mount("/static", StaticFiles(directory="static"), name="static")

# 📌 홈페이지 (이슈 목록 페이지)
@app.get("/")
def home(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # 🔹 이슈 목록 가져오기
    issues = db.query(Issue).all()

    # 🔹 이슈 목록을 템플릿으로 렌더링
    return templates.TemplateResponse("index.html", {
        "request": request,
        "username": user["username"],
        "issues": issues  # 이슈 목록을 템플릿으로 전달
    })
def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


class IssueSchema(BaseModel):
    id: int
    issue_number: str
    issue_name: str
    issue_date: date
    response_period: str
    category: str
    response_team: str
    government_officials: str
    business_impact: str
    kpi: str
    issue_end_date: Optional[date]
    stakeholders: str
    result_summary: str
    completion_status: str
    other_remarks: str
    is_hidden: bool
    updated_at: Optional[date]

    class Config:
        orm_mode = True

# ✅ JSON 응답 API (프론트에서 fetch 등으로 호출할 때 사용)
@app.get("/api/issues/{issue_id}")
def get_issue(issue_id: int, db: Session = Depends(get_db)):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="이슈를 찾을 수 없습니다.")
    return {"message": "이슈 상세", "data": issue}


# ✅ HTML 페이지 렌더링용 (템플릿 응답)
@app.get("/issue/{issue_id}")
def issue_detail_page(
    request: Request,
    issue_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user_required)
):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        return RedirectResponse("/issue-not-found", status_code=302)

    # 🔒 열람 권한 체크
    if user["role"] != "admin":
        if issue.authorized_users:
            authorized_list = [u.strip() for u in issue.authorized_users.split(",") if u.strip()]
            if user["username"] not in authorized_list:
                raise HTTPException(status_code=403, detail="해당 이슈에 대한 열람 권한이 없습니다.")
        elif issue.is_hidden:
            raise HTTPException(status_code=403, detail="열람 권한이 없습니다.")

    details = db.query(IssueDetail).filter(IssueDetail.issue_id == issue_id).order_by(IssueDetail.date).all()

    return templates.TemplateResponse("issue_detail.html", {
        "request": request,
        "issue_id": issue_id,
        "issue": issue,
        "details": details
    })

# 📌 이슈 입력 페이지
@app.get("/create-issue")
def create_issue_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("create_issue.html", {
        "request": request,
        "issue": None,
        "USERS": USERS
    })

@app.post("/create-issue")
async def create_issue_post(
    request: Request,
    issue_number: str = Form(...),
    issue_name: str = Form(...),
    issue_date: date = Form(...),
    response_period: str = Form(...),
    category: str = Form(...),
    response_team: str = Form(...),
    government_officials: str = Form(...),
    business_impact: str = Form(...),
    kpi: str = Form(...),
    issue_end_date: str = Form(...),
    stakeholders: str = Form(...),
    summary: Optional[str] = Form(None),
    completion_status: str = Form(...),
    remarks: Optional[str] = Form(None),
    is_hidden: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    form = await request.form()
    authorized_users = form.getlist("authorized_users")
    authorized_users_str = "," + ",".join(authorized_users) + "," if authorized_users else ""

    # 종료일 파싱
    if issue_end_date.strip() != "미정":
        try:
            parsed_date = datetime.strptime(issue_end_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="종료일 형식이 잘못되었습니다.")
    else:
        parsed_date = None

    new_issue = Issue(
        issue_number=issue_number,
        issue_name=issue_name,
        issue_date=issue_date,
        response_period=response_period,
        category=category,
        response_team=response_team,
        government_officials=government_officials,
        business_impact=business_impact,
        kpi=kpi,
        issue_end_date=parsed_date,
        stakeholders=stakeholders,
        result_summary=summary,
        completion_status=completion_status,
        other_remarks=remarks,
        is_hidden=is_hidden == "on",
        authorized_users=authorized_users_str,
        updated_at=date.today()
    )

    db.add(new_issue)
    db.commit()
    db.refresh(new_issue)

    return RedirectResponse(url="/issues/", status_code=303)

# 📌 이슈 수정 페이지
@app.get("/edit-issue/{issue_id}")
def edit_issue_page(
    request: Request,
    issue_id: int,
    user: dict = Depends(get_current_user_required),
    db: Session = Depends(get_db)
):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        return RedirectResponse("/issue-not-found", status_code=302)
    
    return templates.TemplateResponse("edit_issue.html", {
        "request": request,
        "issue": issue,
        "issue_id": issue_id,
        "USERS": USERS
    })

# 📌 이슈 수정 API
class IssueUpdate(BaseModel):
    issue_number: str
    issue_name: str
    issue_date: date
    response_period: str
    category: str
    response_team: str
    government_officials: str
    business_impact: str
    kpi: str
    issue_end_date: Optional[str]
    stakeholders: str
    result_summary: str
    completion_status: str
    is_hidden: bool  # ✅ 반드시 포함
    other_remarks: str
    authorized_users: Optional[List[str]] = None

@app.post("/edit-issue/{issue_id}")
async def update_issue(
    request: Request,
    issue_id: int,
    user: dict = Depends(get_current_user_required),
    db: Session = Depends(get_db),
    issue_number: str = Form(...),
    issue_name: str = Form(...),
    issue_date: date = Form(...),
    response_period: str = Form(...),
    category: str = Form(...),
    response_team: str = Form(...),
    government_officials: str = Form(...),
    business_impact: str = Form(...),
    kpi: str = Form(...),
    issue_end_date: date = Form(...),
    stakeholders: str = Form(...),
    summary: str = Form(...),
    completion_status: str = Form(...),
    remarks: str = Form(...),
    is_hidden: Optional[str] = Form(None)
):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        return RedirectResponse("/issue-not-found", status_code=302)

    # 기본 필드 업데이트
    issue.issue_number = issue_number
    issue.issue_name = issue_name
    issue.issue_date = issue_date
    issue.response_period = response_period
    issue.category = category
    issue.response_team = response_team
    issue.government_officials = government_officials
    issue.business_impact = business_impact
    issue.kpi = kpi
    issue.issue_end_date = issue_end_date
    issue.stakeholders = stakeholders
    issue.summary = summary
    issue.completion_status = completion_status
    issue.remarks = remarks
    issue.is_hidden = is_hidden == "on"

    # 🔥 여기서 authorized_users 추출
    form = await request.form()
    authorized_users = form.getlist("authorized_users")
    issue.authorized_users = "," + ",".join(authorized_users) + "," if authorized_users else ""

    db.commit()
    return RedirectResponse("/issue-list", status_code=302)

@app.put("/issues/{issue_id}")
def update_issue(
    issue_id: int,
    issue_update: IssueUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="이슈를 찾을 수 없습니다.")

    # ✅ 종료일 처리
    if issue_update.issue_end_date and issue_update.issue_end_date.strip() != "미정":
        try:
            parsed_date = datetime.strptime(issue_update.issue_end_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="이슈 종료일 형식이 잘못되었습니다.")
    else:
        parsed_date = None

    # ✅ authorized_users 리스트를 문자열로 변환해서 저장
    issue_data = issue_update.dict(exclude={"issue_end_date"})
    if issue_update.authorized_users:
        issue_data["authorized_users"] = ",".join(issue_update.authorized_users)
    else:
        issue_data["authorized_users"] = ""

    for key, value in issue_data.items():
        setattr(issue, key, value)

    issue.issue_end_date = parsed_date
    issue.updated_at = date.today()

    db.commit()
    db.refresh(issue)

    return {"message": "이슈가 수정되었습니다.", "data": issue}

# 📌 세부 진행 사항 추가 페이지
@app.get("/issue/{issue_id}/add-detail")
def add_issue_detail_page(request: Request, issue_id: int, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("add_or_edit_issue_detail.html", {
        "request": request,
        "issue_id": issue_id,
        "mode": "추가",
        "detail": None
    })

@app.get("/issue/detail-edit/{detail_id}")
def edit_issue_detail_page(request: Request, detail_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    detail = db.query(IssueDetail).filter(IssueDetail.id == detail_id).first()
    if not detail:
        raise HTTPException(status_code=404, detail="진행 사항을 찾을 수 없습니다.")
    return templates.TemplateResponse("add_or_edit_issue_detail.html", {
        "request": request,
        "mode": "수정",
        "detail": detail
    })

@app.get("/download-excel/")
def download_excel(db: Session = Depends(get_db)):
    issues = db.query(Issue).all()

    if not issues:
        raise HTTPException(status_code=404, detail="다운로드할 이슈가 없습니다.")

    # ✅ 데이터프레임으로 변환
    data = [
        {
            "이슈 번호": issue.issue_number,
            "이슈명": issue.issue_name,
            "발생일": issue.issue_date,
            "대응 기간": issue.response_period,
            "카테고리": issue.category,
            "대응팀": issue.response_team,
            "정부 관계자": issue.government_officials,
            "Business Impact": issue.business_impact,
            "KPI": issue.kpi,
            "이슈 종료일(예정일)": issue.issue_end_date,
            "이해관계자": issue.stakeholders,
            "결과 요약": issue.result_summary,
            "완료 여부": issue.completion_status,
            "기타 사항": issue.other_remarks
        }
        for issue in issues
    ]
    df = pd.DataFrame(data)

    # ✅ 엑셀 파일 저장 경로 설정
    file_path = "issues.xlsx"
    df.to_excel(file_path, index=False, engine="openpyxl")

    return FileResponse(file_path, filename="이슈목록.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/issues/{issue_id}/details/download", response_class=StreamingResponse)
def download_issue_details(
    issue_id: int = Path(..., description="Issue ID in URL Path"), 
    db: Session = Depends(get_db)
):
    print(f"\n[DEBUG] Received issue_id: {issue_id} (Type: {type(issue_id)})")

    # ✅ 데이터베이스에서 이슈 정보 가져오기
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        print(f"[ERROR] Issue {issue_id} not found in database")
        raise HTTPException(status_code=404, detail="Issue not found")

    # ✅ 세부 진행 사항 가져오기
    details = db.query(IssueDetail).filter(IssueDetail.issue_id == issue_id).order_by(IssueDetail.date.asc()).all()
    if not details:
        print(f"[ERROR] No details found for issue_id {issue_id}")
        raise HTTPException(status_code=404, detail="No details found for this issue")

    print(f"[DEBUG] Found {len(details)} details for issue_id {issue_id}")

    # ✅ 데이터를 DataFrame으로 변환 (첨부파일 제거)
    df = pd.DataFrame([{
        "날짜": detail.date,
        "내용": detail.content
    } for detail in details])

    # ✅ 엑셀 파일을 메모리에 저장
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="진행사항")
    output.seek(0)

    # ✅ 파일명: "이슈명 진행사항.xlsx" 형식으로 변경
    issue_name = issue.issue_name  # 이슈명 가져오기
    safe_issue_name = issue_name.replace(" ", "_")  # 파일명 안전하게 변경 (공백 -> _)
    filename = f"{safe_issue_name}_진행사항.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )

# 📌 Excel 업로드 API
@app.post("/upload-excel/")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    # ✅ 업로드된 파일 확장자 확인
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀 파일 (.xlsx)만 업로드 가능합니다.")

    # ✅ 파일을 메모리에서 읽기
    contents = file.file.read()
    df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")

    # ✅ 컬럼명 검증
    expected_columns = [
        "이슈 번호", "이슈명", "발생일", "대응 기간", "카테고리", "대응팀",
        "정부 관계자", "Business Impact", "KPI", "이슈 종료일(예정일)", "이해관계자",
        "결과 요약", "완료 여부", "기타 사항"
    ]
    
    if not all(col in df.columns for col in expected_columns):
        raise HTTPException(status_code=400, detail="엑셀 컬럼명이 올바르지 않습니다.")

    # ✅ 데이터베이스에 저장
    for _, row in df.iterrows():
        new_issue = Issue(
            issue_number=row["이슈 번호"],
            issue_name=row["이슈명"],
            issue_date=row["발생일"],
            response_period=row["대응 기간"],
            category=row["카테고리"],
            response_team=row["대응팀"],
            government_officials=row["정부 관계자"],
            business_impact=row["Business Impact"],
            kpi=None if pd.isna(row["KPI"]) else row["KPI"],
            issue_end_date=None if pd.isna(row["이슈 종료일(예정일)"]) else row["이슈 종료일(예정일)"],
            stakeholders=row["이해관계자"],
            result_summary=None if pd.isna(row["결과 요약"]) else row["결과 요약"],
            completion_status=row["완료 여부"],
            other_remarks=None if pd.isna(row["기타 사항"]) else row["기타 사항"],
            updated_at=date.today()
        )
        db.add(new_issue)

    db.commit()
    return {"message": f"{len(df)}개의 이슈가 성공적으로 추가되었습니다."}

@app.get("/upload-excel")
def upload_excel_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("upload_excel.html", {"request": request})

# 📌 Pydantic 모델 (입력 데이터 검증)
class ReviewProposalCreate(BaseModel):
    content: str
    reason: str
    goal: str

# 📌 현안 검토 건의 페이지 렌더링
@app.get("/review", response_class=HTMLResponse)
async def review_page(request: Request, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    user = request.session.get("user")
    reviews_raw = db.query(ReviewProposal).order_by(ReviewProposal.date.desc()).all()

    reviews = [{
        "id": r.id,
        "date": r.date.strftime("%Y-%m-%d"),
        "content": r.content,
        "reason": r.reason,
        "goal": r.goal
    } for r in reviews_raw]

    return templates.TemplateResponse("review.html", {
        "request": request,
        "reviews": reviews,
        "user": user
    })

# 📌 검토 건의 추가 API
@app.post("/review/add")
def add_review(
    content: str = Form(...),
    reason: str = Form(...),
    goal: str = Form(...), 
    date: str = Form(None),  # `date`는 문자열로 들어옴
    db: Session = Depends(get_db)
):
    if not date:  
        date = datetime.today().date()  # ✅ 오늘 날짜 자동 입력 (datetime.date 객체)
    else:
        date = datetime.strptime(date, "%Y-%m-%d").date()  # ✅ 문자열 → datetime.date 변환

    new_review = ReviewProposal(
        date=date,
        content=content,
        reason=reason,
        goal=goal
    )
    db.add(new_review)
    db.commit()
    return RedirectResponse(url="/review", status_code=302)

# 📌 검토 건의 삭제 API
@app.post("/review/delete/{review_id}")
async def delete_review(review_id: int, db: Session = Depends(get_db)):
    review = db.query(ReviewProposal).filter(ReviewProposal.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="검토 건의를 찾을 수 없습니다.")

    db.delete(review)
    db.commit()
    return RedirectResponse(url="/review", status_code=303)

# 📌 세부 진행 사항 엑셀 업로드 API
@app.post("/issues/{issue_id}/details/upload-excel/")
def upload_issue_details_excel(
    issue_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    # ✅ 업로드된 파일 확장자 확인
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀 파일 (.xlsx)만 업로드 가능합니다.")

    # ✅ 파일을 메모리에서 읽기
    contents = file.file.read()
    df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")

    # ✅ 컬럼명 검증
    expected_columns = ["날짜", "내용"]
    if not all(col in df.columns for col in expected_columns):
        raise HTTPException(status_code=400, detail="엑셀 컬럼명이 올바르지 않습니다. ('날짜', '내용')")

    # ✅ 데이터베이스에 저장
    for _, row in df.iterrows():
        new_detail = IssueDetail(
            issue_id=issue_id,
            date=row["날짜"],
            content=row["내용"],
            file_path="[]"  # 파일 없음
        )
        db.add(new_detail)

    db.commit()
    return {"message": f"{len(df)}개의 세부 진행 사항이 성공적으로 추가되었습니다."}

@app.get("/law-update", response_class=HTMLResponse)
async def get_law_updates(request: Request, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    updates = db.query(LawUpdate).order_by(LawUpdate.proclamation_date.desc()).all()
    user = request.session.get("user", {"role": "user"})

    # ✅ file_path를 미리 JSON으로 파싱해서 넘겨주기
    processed_updates = []
    for update in updates:
        try:
            file_list = json.loads(update.file_path) if update.file_path else []
        except json.JSONDecodeError:
            file_list = []

        processed_updates.append({
            "id": update.id,
            "category": update.category,
            "law_name": update.law_name,
            "proclamation_date": update.proclamation_date,
            "content": update.content,
            "file_list": file_list,
            "notice_date": update.notice_date  # ✅ 추가
        })

    return templates.TemplateResponse("law_update.html", {
        "request": request,
        "updates": processed_updates,
        "user": user
    })

# 법령 개정 추가 API
@app.post("/law-update/add")
async def add_law_update(  # ✅ 반드시 async def 여야 함
    category: str = Form(...),
    law_name: str = Form(...),
    proclamation_date: date = Form(...),
    content: str = Form(...),
    related_files: List[UploadFile] = File(None),
    notice_date: Optional[date] = Form(None),
    db: Session = Depends(get_db)
):
    file_paths = []
    if related_files:
        for file in related_files:
            if file.filename:
                folder_name = f"{law_name}_{proclamation_date.strftime('%Y%m%d')}"
                folder_path = os.path.join(UPLOAD_DIR, folder_name)
                os.makedirs(folder_path, exist_ok=True)

                file_name = file.filename
                full_path = os.path.join(folder_path, file_name)

                file_content = await file.read()  # ✅ 이건 async def 안에서만 가능!
                with open(full_path, "wb") as f:
                    f.write(file_content)

                file_paths.append(f"{folder_name}/{file_name}")

    new_law = LawUpdate(
        category=category,
        law_name=law_name,
        proclamation_date=proclamation_date,
        content=content,
        file_path=json.dumps(file_paths) if file_paths else None,
        notice_date=notice_date  # ✅ 추가
    )
    db.add(new_law)
    db.commit()
    return RedirectResponse(url="/law-update", status_code=303)

# 법령 개정 삭제 API
@app.post("/law-update/delete/{law_id}")
async def delete_law_update(law_id: int, db: Session = Depends(get_db)):
    law = db.query(LawUpdate).filter(LawUpdate.id == law_id).first()
    if not law:
        raise HTTPException(status_code=404, detail="법령 개정을 찾을 수 없습니다.")
    db.delete(law)
    db.commit()
    return RedirectResponse(url="/law-update", status_code=303)

@app.get("/law-update-data")
def get_law_updates(db: Session = Depends(get_db)):
    laws = db.query(LawUpdate).order_by(LawUpdate.proclamation_date.desc()).all()  # ✅ 여기 수정
    return [
        {
            "id": law.id,
            "category": law.category,
            "law_name": law.law_name,
            "proclamation_date": law.proclamation_date.strftime("%Y-%m-%d"),  # ✅ 여기 이름도 통일
            "content": law.content or "",
            "file_list": json.loads(law.file_path) if law.file_path else [],
            "notice_date": law.notice_date.strftime("%Y-%m-%d") if law.notice_date else ""
        } for law in laws
    ]

@app.post("/law-update/edit/{law_id}")
async def update_law(
    law_id: int,
    category: str = Form(...),
    law_name: str = Form(...),
    proclamation_date: date = Form(...),
    notice_date: Optional[date] = Form(None),
    content: str = Form(...),
    db: Session = Depends(get_db)
):
    law = db.query(LawUpdate).filter(LawUpdate.id == law_id).first()
    if not law:
        raise HTTPException(status_code=404, detail="법령 항목을 찾을 수 없습니다.")

    law.category = category
    law.law_name = law_name
    law.proclamation_date = proclamation_date
    law.notice_date = notice_date
    law.content = content

    db.commit()
    return RedirectResponse(url="/law-update", status_code=303)

@app.get("/law-update/add-form", response_class=HTMLResponse)
def law_add_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("law_update_add.html", {"request": request})


@app.get("/law-update/edit/{law_id}", response_class=HTMLResponse)
def law_edit_page(request: Request, law_id: int, db: Session = Depends(get_db)):
    law = db.query(LawUpdate).filter(LawUpdate.id == law_id).first()
    if not law:
        raise HTTPException(status_code=404, detail="해당 항목을 찾을 수 없습니다.")
    return templates.TemplateResponse("law_update_edit.html", {"request": request, "law": law})

@app.get("/files/{file_name}")
async def download_file(file_name: str):
    file_path = os.path.join(UPLOAD_DIR, file_name)
    return FileResponse(file_path, media_type="application/octet-stream", filename=file_name)

@app.get("/wg-operation", response_class=HTMLResponse)
async def wg_operation_menu(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("wg_menu.html", {"request": request})

@app.get("/wg/physicochemical", response_class=HTMLResponse)
async def wg_physicochemical_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("physicochemical.html", {"request": request})

@app.get("/wg/efficacy", response_class=HTMLResponse)
async def wg_efficacy_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("efficacy.html", {"request": request})

@app.get("/wg/residue", response_class=HTMLResponse)
async def wg_residue_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("residue.html", {"request": request})

@app.get("/wg/toxicity", response_class=HTMLResponse)
async def wg_toxicity_page(request: Request, user: dict = Depends(get_current_user_required)):
    return templates.TemplateResponse("toxicity.html", {"request": request})

@app.post("/physicochemical/add")
async def add_physicochemical(
    category: str = Form(...),
    date: date = Form(...),
    subject: str = Form(...),
    content: str = Form(...),
    files: List[UploadFile] = File(None),
    db: Session = Depends(get_db),
    request: Request = None
):
    file_paths = []
    if files:
        for file in files:
            if file.filename:
                folder_path = os.path.join(UPLOAD_DIR, "physicochemical")
                os.makedirs(folder_path, exist_ok=True)

                full_path = os.path.join(folder_path, file.filename)
                with open(full_path, "wb") as f:
                    f.write(await file.read())

                file_paths.append(f"physicochemical/{file.filename}")

    new_item = Physicochemical(
        category=category,
        date=date,
        subject=subject,
        content=content,
        file_path=json.dumps(file_paths)
    )

    db.add(new_item)
    db.commit()
    return RedirectResponse(url="/wg/physicochemical", status_code=303)

@app.get("/physicochemical/list")
def list_physicochemical(db: Session = Depends(get_db)):
    items = db.query(Physicochemical).order_by(Physicochemical.date.desc()).all()
    result = []
    for item in items:
        file_list = json.loads(item.file_path) if item.file_path else []
        result.append({
            "id": item.id,
            "category": item.category,
            "date": item.date.strftime("%Y-%m-%d"),
            "subject": item.subject,
            "content": item.content,
            "file_list": file_list
        })
    return result

@app.post("/physicochemical/delete/{item_id}")
async def delete_physicochemical(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Physicochemical).filter(Physicochemical.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/physicochemical", status_code=303)

@app.post("/efficacy/add")
async def add_efficacy(
    category: str = Form(...),
    date: date = Form(...),
    subject: str = Form(...),
    content: str = Form(...),
    files: List[UploadFile] = File(None),
    db: Session = Depends(get_db),
    request: Request = None
):
    file_paths = []
    if files:
        for file in files:
            if file.filename:
                folder_path = os.path.join(UPLOAD_DIR, "efficacy")
                os.makedirs(folder_path, exist_ok=True)

                full_path = os.path.join(folder_path, file.filename)
                with open(full_path, "wb") as f:
                    f.write(await file.read())

                file_paths.append(f"efficacy/{file.filename}")

    new_item = Efficacy(
        category=category,
        date=date,
        subject=subject,
        content=content,
        file_path=json.dumps(file_paths)
    )

    db.add(new_item)
    db.commit()
    return RedirectResponse(url="/wg/efficacy", status_code=303)

@app.get("/efficacy/list")
def list_efficacy(db: Session = Depends(get_db)):
    items = db.query(Efficacy).order_by(Efficacy.date.desc()).all()
    result = []
    for item in items:
        file_list = json.loads(item.file_path) if item.file_path else []
        result.append({
            "id": item.id,
            "category": item.category,
            "date": item.date.strftime("%Y-%m-%d"),
            "subject": item.subject,
            "content": item.content,
            "file_list": file_list
        })
    return result

@app.post("/efficacy/delete/{item_id}")
async def delete_efficacy(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Efficacy).filter(Efficacy.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/efficacy", status_code=303)

@app.post("/physicochemical/delete/{item_id}")
async def delete_physicochemical(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Physicochemical).filter(Physicochemical.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/physicochemical", status_code=303)

@app.post("/residue/add")
async def add_residue(
    category: str = Form(...),
    date: date = Form(...),
    subject: str = Form(...),
    content: str = Form(...),
    files: List[UploadFile] = File(None),
    db: Session = Depends(get_db),
    request: Request = None
):
    file_paths = []
    if files:
        for file in files:
            if file.filename:
                folder_path = os.path.join(UPLOAD_DIR, "residue")
                os.makedirs(folder_path, exist_ok=True)

                full_path = os.path.join(folder_path, file.filename)
                with open(full_path, "wb") as f:
                    f.write(await file.read())

                file_paths.append(f"residue/{file.filename}")

    new_item = Residue(
        category=category,
        date=date,
        subject=subject,
        content=content,
        file_path=json.dumps(file_paths)
    )

    db.add(new_item)
    db.commit()
    return RedirectResponse(url="/wg/residue", status_code=303)

@app.get("/residue/list")
def list_residue(db: Session = Depends(get_db)):
    items = db.query(Residue).order_by(Residue.date.desc()).all()
    result = []
    for item in items:
        file_list = json.loads(item.file_path) if item.file_path else []
        result.append({
            "id": item.id,
            "category": item.category,
            "date": item.date.strftime("%Y-%m-%d"),
            "subject": item.subject,
            "content": item.content,
            "file_list": file_list
        })
    return result

@app.post("/residue/delete/{item_id}")
async def delete_residue(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Residue).filter(Residue.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/residue", status_code=303)

@app.post("/physicochemical/delete/{item_id}")
async def delete_physicochemical(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Physicochemical).filter(Physicochemical.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/physicochemical", status_code=303)

@app.post("/toxicity/add")
async def add_toxicity(
    category: str = Form(...),
    date: date = Form(...),
    subject: str = Form(...),
    content: str = Form(...),
    files: List[UploadFile] = File(None),
    db: Session = Depends(get_db),
    request: Request = None
):
    file_paths = []
    if files:
        for file in files:
            if file.filename:
                folder_path = os.path.join(UPLOAD_DIR, "toxicity")
                os.makedirs(folder_path, exist_ok=True)

                full_path = os.path.join(folder_path, file.filename)
                with open(full_path, "wb") as f:
                    f.write(await file.read())

                file_paths.append(f"toxicity/{file.filename}")

    new_item = Toxicity(
        category=category,
        date=date,
        subject=subject,
        content=content,
        file_path=json.dumps(file_paths)
    )

    db.add(new_item)
    db.commit()
    return RedirectResponse(url="/wg/toxicity", status_code=303)

@app.get("/toxicity/list")
def list_toxicity(db: Session = Depends(get_db)):
    items = db.query(Toxicity).order_by(Toxicity.date.desc()).all()
    result = []
    for item in items:
        file_list = json.loads(item.file_path) if item.file_path else []
        result.append({
            "id": item.id,
            "category": item.category,
            "date": item.date.strftime("%Y-%m-%d"),
            "subject": item.subject,
            "content": item.content,
            "file_list": file_list
        })
    return result

@app.post("/toxicity/delete/{item_id}")
async def delete_toxicity(item_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    item = db.query(Toxicity).filter(Toxicity.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="데이터를 찾을 수 없습니다.")
    
    db.delete(item)
    db.commit()
    return RedirectResponse(url="/wg/toxicity", status_code=303)

@app.post("/supplements/add")
async def add_supplement(
    supplement_date: date = Form(...),
    category: str = Form(...),
    content: str = Form(...),
    responder: str = Form(...),
    related_files: List[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    file_paths = []
    if related_files:
        folder_path = os.path.join(UPLOAD_DIR, "supplements")
        os.makedirs(folder_path, exist_ok=True)

        for file in related_files:
            if file.filename:
                file_path = os.path.join(folder_path, file.filename)
                with open(file_path, "wb") as f:
                    f.write(await file.read())
                file_paths.append(f"supplements/{file.filename}")

    new_item = Supplement(
        supplement_date=supplement_date,
        category=category,
        content=content,
        responder=responder,
        file_path=json.dumps(file_paths)
    )

    db.add(new_item)
    db.commit()
    return RedirectResponse(url="/supplements", status_code=303)

@app.get("/supplements", response_class=HTMLResponse)
async def list_supplements(
    request: Request,
    db: Session = Depends(get_db),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    content: Optional[str] = Query(None),
    responder: Optional[str] = Query(None),
    user: dict = Depends(get_current_user_required),
):
    query = db.query(Supplement)
    if start_date:
        query = query.filter(Supplement.supplement_date >= start_date)
    if end_date:
        query = query.filter(Supplement.supplement_date <= end_date)
    if category:
        query = query.filter(Supplement.category == category)
    if content:
        query = query.filter(Supplement.content.ilike(f"%{content}%"))
    if responder:
        query = query.filter(Supplement.responder.ilike(f"%{responder}%"))
    query = query.order_by(Supplement.supplement_date.desc())
    items = query.all()
    supplements = []
    for item in items:
        file_list = json.loads(item.file_path) if item.file_path else []
        supplements.append({
            "id": item.id,
            "supplement_date": item.supplement_date,
            "category": item.category,
            "content": item.content,
            "responder": item.responder,
            "file_list": file_list,
            "response_method": item.response_method,
            "response_result": item.response_result
        })
    return templates.TemplateResponse("supplements.html", {
        "request": request,
        "supplements": supplements,
        "user": user
    })

@app.get("/supplements-data")
def get_supplements_json(
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(2, le=100)
):
    total_count = db.query(Supplement).count()
    supplements_raw = db.query(Supplement)\
        .order_by(Supplement.supplement_date.desc())\
        .offset((page - 1) * page_size)\
        .limit(page_size)\
        .all()

    supplements = []
    for item in supplements_raw:
        file_list = json.loads(item.file_path) if item.file_path else []
        supplements.append({
            "id": item.id,
            "supplement_date": item.supplement_date.strftime("%Y-%m-%d"),
            "category": item.category,
            "content": item.content,
            "responder": item.responder,
            "file_list": file_list,
            "response_method": item.response_method,
            "response_result": item.response_result
        })

    return {
        "data": supplements,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": (total_count + page_size - 1) // page_size
    }

@app.put("/supplements/{id}/update-response")
async def update_supplement_response(
    id: int,
    data: SupplementResponseUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    supplement = db.query(Supplement).filter(Supplement.id == id).first()
    if not supplement:
        raise HTTPException(status_code=404, detail="보완사항을 찾을 수 없습니다.")

    supplement.response_method = data.response_method
    supplement.response_result = data.response_result

    db.commit()
    return {"message": "대응방법 및 결과가 저장되었습니다."}

@app.post("/supplements/delete/{id}")
async def delete_supplement(id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    supplement = db.query(Supplement).filter(Supplement.id == id).first()
    if not supplement:
        raise HTTPException(status_code=404, detail="보완사항을 찾을 수 없습니다.")

    db.delete(supplement)
    db.commit()
    return RedirectResponse(url="/supplements", status_code=303)

@app.get("/admin-login")
def admin_login_page(request: Request):
    return templates.TemplateResponse("admin_login.html", {"request": request})

@app.get("/admin-login", response_class=HTMLResponse)
def admin_login_page(request: Request):
    return templates.TemplateResponse("admin_login.html", {"request": request})

@app.post("/admin-login", response_class=HTMLResponse)
def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    user = USERS.get(username)
    if not user or user["password"] != password or user["role"] != "admin":
        return templates.TemplateResponse("admin_login.html", {
            "request": request,
            "error": "관리자 아이디 또는 비밀번호가 잘못되었습니다."
        })

    expires_at = datetime.utcnow() + timedelta(seconds=1800)
    request.session["user"] = {
        "username": username,
        "role": user["role"],
        "expires_at": expires_at.isoformat()
    }
    return RedirectResponse(url="/select-menu", status_code=302)

@app.get("/test-review", response_class=HTMLResponse)
async def test_review(request: Request, user: dict = Depends(get_current_user_required), db: Session = Depends(get_db)):
    test_designs = db.query(TestReview).order_by(TestReview.id.desc()).all()
    return templates.TemplateResponse("test_review.html", {
        "request": request,
        "test_designs": test_designs,
        "user": user
    })

@app.get("/test-review/add", response_class=HTMLResponse)
async def show_add_form(request: Request, user: dict = Depends(get_current_user_required)):
    if request.session.get("user")["role"] != "admin":
        return RedirectResponse("/select-menu", status_code=303)
    return templates.TemplateResponse("test_review_add.html", {"request": request})

@app.post("/test-review/add")
async def add_test_review(
    request: Request,
    year: int = Form(...),
    category: str = Form(...),
    field: str = Form(...),
    crop_name: str = Form(...),
    pest_name: str = Form(...),
    institution_name: str = Form(...),
    review_type: str = Form(...),
    review_item: str = Form(...),
    review_result: str = Form(...),
    db: Session = Depends(get_db)
):
    new_review = TestReview(
        year=year,
        category=category,
        field=field,
        crop_name=crop_name,
        pest_name=pest_name,
        institution_name=institution_name,
        review_type=review_type,
        review_item=review_item,
        review_result=review_result
    )
    db.add(new_review)
    db.commit()
    db.refresh(new_review)

    return RedirectResponse("/test-review", status_code=303)

Base.metadata.create_all(bind=engine)

# ✅ authorized_users 쉼표 감싸기 자동 처리 (서버 최초 실행 시 1회만 적용)
def fix_authorized_users_format():
    db: Session = SessionLocal()
    try:
        issues = db.query(Issue).filter(Issue.authorized_users != None).all()
        for issue in issues:
            if issue.authorized_users and not issue.authorized_users.startswith(","):
                issue.authorized_users = "," + issue.authorized_users.strip(",") + ","
        db.commit()
    finally:
        db.close()

fix_authorized_users_format()  # 🚨 서버 시작 시 1회만 적용되게 하세요

@app.get("/test-review/download")
def download_test_review_excel(db: Session = Depends(get_db)):
    data = db.query(TestReview).order_by(TestReview.id).all()

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "시험설계 목록"

    # 헤더
    headers = ["시험년도", "구분", "분야", "작물명", "병해충명", "시험기관명", "검토구분", "검토사항", "검토결과"]
    ws.append(headers)

    # 데이터 추가
    for item in data:
        ws.append([
            item.year,
            item.category,
            item.field,
            item.crop_name,
            item.pest_name,
            item.institution_name,
            item.review_type,
            item.review_item or "",
            item.review_result or ""
        ])

    # 엑셀 메모리에 저장
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_review.xlsx"}
    )

@app.post("/test-review/upload")
async def upload_test_review_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user_required)
):
    if user["role"] != "admin":
        return RedirectResponse("/select-menu", status_code=303)

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if not any(row):
                continue  # skip empty row

            new_entry = TestReview(
                year=int(row[0]),
                category=row[1],
                field=row[2],
                crop_name=row[3],
                pest_name=row[4],
                institution_name=row[5],
                review_type=row[6],
                review_item=row[7] or "",
                review_result=row[8] or ""
            )
            db.add(new_entry)

        db.commit()
        return RedirectResponse("/test-review", status_code=303)

    except Exception as e:
        print("엑셀 업로드 중 오류:", e)
        return HTMLResponse(content=f"<h3>엑셀 파일 업로드 실패: {str(e)}</h3>", status_code=400)

@app.get("/test-review/data")
async def get_filtered_test_reviews(
    year: Optional[int] = Query(None),
    category: Optional[str] = Query(None),
    field: Optional[str] = Query(None),
    crop_name: Optional[str] = Query(None),
    pest_name: Optional[str] = Query(None),
    institution_name: Optional[str] = Query(None),
    review_type: Optional[str] = Query(None),
    review_item: Optional[str] = Query(None),
    review_result: Optional[str] = Query(None),
    page: int = Query(1, ge=1),  # 페이지 번호 (기본값 1)
    page_size: int = Query(10, le=100),  # 페이지당 항목 수 (기본값 10, 최대 100)
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user_required)
):
    query = db.query(TestReview)

    if year:
        query = query.filter(TestReview.year == year)
    if category:
        query = query.filter(TestReview.category.ilike(f"%{category}%"))
    if field:
        query = query.filter(TestReview.field.ilike(f"%{field}%"))
    if crop_name:
        query = query.filter(TestReview.crop_name.ilike(f"%{crop_name}%"))
    if pest_name:
        query = query.filter(TestReview.pest_name.ilike(f"%{pest_name}%"))
    if institution_name:
        query = query.filter(TestReview.institution_name.ilike(f"%{institution_name}%"))
    if review_type:
        query = query.filter(TestReview.review_type.ilike(f"%{review_type}%"))
    if review_item:
        query = query.filter(TestReview.review_item.ilike(f"%{review_item}%"))
    if review_result:
        query = query.filter(TestReview.review_result.ilike(f"%{review_result}%"))

    # Total count of records for pagination
    total_count = query.count()

    # Apply pagination
    test_reviews = query.offset((page - 1) * page_size).limit(page_size).all()

    # Format the results
    results = [
        {
            "id": review.id,
            "year": review.year,
            "category": review.category,
            "field": review.field,
            "crop_name": review.crop_name,
            "pest_name": review.pest_name,
            "institution_name": review.institution_name,
            "review_type": review.review_type,
            "review_item": review.review_item,
            "review_result": review.review_result,
        }
        for review in test_reviews
    ]

    # Return data with pagination information
    return {
        "data": results,
        "user_role": user["role"],
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": (total_count + page_size - 1) // page_size,  # Calculate total pages
    }

# 수정 라우트
@app.get("/test-review/edit/{id}", response_class=HTMLResponse)
async def edit_test_review(id: int, request: Request, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="권한이 없습니다.")

    test_design = db.query(TestReview).filter(TestReview.id == id).first()
    if not test_design:
        raise HTTPException(status_code=404, detail="시험설계 정보를 찾을 수 없습니다.")

    return templates.TemplateResponse("test_review_edit.html", {"request": request, "test_design": test_design})

# 수정된 데이터 저장
@app.post("/test-review/edit/{id}")
async def save_test_review_edit(id: int, request: Request, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="권한이 없습니다.")

    data = await request.form()
    year = data.get("year")
    category = data.get("category")
    field = data.get("field")
    crop_name = data.get("crop_name")
    pest_name = data.get("pest_name")
    institution_name = data.get("institution_name")
    review_type = data.get("review_type")
    review_item = data.get("review_item")
    review_result = data.get("review_result")

    test_design = db.query(TestReview).filter(TestReview.id == id).first()
    if not test_design:
        raise HTTPException(status_code=404, detail="시험설계 정보를 찾을 수 없습니다.")
    
    test_design.year = year
    test_design.category = category
    test_design.field = field
    test_design.crop_name = crop_name
    test_design.pest_name = pest_name
    test_design.institution_name = institution_name
    test_design.review_type = review_type
    test_design.review_item = review_item
    test_design.review_result = review_result

    db.commit()
    return RedirectResponse("/test-review", status_code=303)

# 삭제 라우트
@app.post("/test-review/delete/{id}")
async def delete_test_review(id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user_required)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="권한이 없습니다.")

    test_design = db.query(TestReview).filter(TestReview.id == id).first()
    if not test_design:
        raise HTTPException(status_code=404, detail="시험설계 정보를 찾을 수 없습니다.")

    db.delete(test_design)
    db.commit()
    
    return RedirectResponse("/test-review", status_code=303)

@app.get("/supplements/add", response_class=HTMLResponse)
async def supplement_add_form(request: Request, user: dict = Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403)
    return templates.TemplateResponse("add_supplement.html", {"request": request})

@app.post("/supplements/upload")
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    upload_dir = "uploads/"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            raw_date = row[0].value
            if isinstance(raw_date, str):
                supplement_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
            elif isinstance(raw_date, datetime):
                supplement_date = raw_date.date()
            else:
                raise ValueError(f"{i}행: 날짜 형식 오류")

            category = row[1].value
            content = row[2].value
            responder = row[3].value

            new_item = Supplement(
                supplement_date=supplement_date,
                category=category,
                content=content,
                responder=responder
            )
            db.add(new_item)

        except Exception as e:
            print(f"{i}행 처리 중 오류:", e)
            continue

    db.commit()
    return RedirectResponse(url="/supplements", status_code=303)

@app.get("/supplements-all-data")
def get_all_supplements_json(db: Session = Depends(get_db)):
    supplements_raw = db.query(Supplement).order_by(Supplement.supplement_date.desc()).all()

    supplements = []
    for item in supplements_raw:
        file_list = json.loads(item.file_path) if item.file_path else []
        supplements.append({
            "id": item.id,
            "supplement_date": item.supplement_date.strftime("%Y-%m-%d"),
            "category": item.category,
            "content": item.content,
            "responder": item.responder,
            "file_list": file_list,
            "response_method": item.response_method,
            "response_result": item.response_result
        })

    return { "data": supplements }

@app.get("/api/stats/summary")
async def get_stats_summary(db: Session = Depends(get_db)):
    total = db.query(Issue).count()
    completed = db.query(Issue).filter(Issue.completion_status == "완료").count()
    in_progress = db.query(Issue).filter(Issue.completion_status == "진행중").count()

    category_count = {}
    for row in db.query(Issue.category).all():
        category_count[row[0]] = category_count.get(row[0], 0) + 1

    impact_count = {"High": 0, "Medium": 0, "Low": 0}
    for row in db.query(Issue.business_impact).all():
        if row[0] in impact_count:
            impact_count[row[0]] += 1

    year_counts_raw = db.query(
        extract('year', Issue.issue_date).label("year"),
        func.count(Issue.id)
    ).group_by("year").all()

    year_counts = {str(int(year)): count for year, count in year_counts_raw}

    return {  # ✅ 여기 들여쓰기 주의!
        "total": total,
        "completed": completed,
        "in_progress": in_progress,
        "category_count": category_count,
        "impact_count": impact_count,
        "year_counts": year_counts,
    }

@app.get("/dashboard", response_class=HTMLResponse)
async def show_dashboard(request: Request):
    return templates.TemplateResponse("dashboard.html", {"request": request})

@app.get("/api/stats/proposals")
async def get_proposal_stats(db: Session = Depends(get_db)):
    year_counts_raw = db.query(
        extract('year', ReviewProposal.date),  # ✅ 정확한 필드명 사용
        func.count()
    ).group_by(extract('year', ReviewProposal.date)).all()

    year_counts = {str(year): count for year, count in year_counts_raw}
    return {"year_counts": year_counts}

@app.get("/api/stats/wg")
async def get_wg_stats(db: Session = Depends(get_db)):
    models = {
        "이화학": Physicochemical,
        "약효약해": Efficacy,
        "잔류성": Residue,
        "독성": Toxicity
    }

    result = {}

    for label, model in models.items():
        year_counts = db.query(
            extract('year', model.date).label("year"),
            func.count().label("count")
        ).group_by("year").all()
        
        # 변환: {'2023': 5, '2024': 7} 형태
        result[label] = {str(int(year)): count for year, count in year_counts}

    return result

@app.get("/api/stats/supplements")
async def get_supplement_stats(db: Session = Depends(get_db)):
    # 연도별 보완 갯수
    year_counts_raw = db.query(
        extract('year', Supplement.supplement_date).label("year"),
        func.count().label("count")
    ).group_by("year").all()

    year_counts = {str(int(year)): count for year, count in year_counts_raw}

    # 보완 분야별 갯수
    category_counts_raw = db.query(Supplement.category, func.count()).group_by(Supplement.category).all()
    category_counts = {category: count for category, count in category_counts_raw}

    return {
        "year_counts": year_counts,
        "category_counts": category_counts
    }

@app.get("/api/stats/test-reviews")
async def get_test_review_stats(db: Session = Depends(get_db)):
    # ✅ 분야별 통계 (예: 살균제, 살충제, 제초제 등)
    field_counts_raw = db.query(TestReview.field, func.count()).group_by(TestReview.field).all()
    field_counts = {field: count for field, count in field_counts_raw}

    # ✅ 연도별 통계
    year_counts_raw = db.query(TestReview.year, func.count()).group_by(TestReview.year).all()
    year_counts = {str(year): count for year, count in year_counts_raw}

    return {
        "field_counts": field_counts,
        "year_counts": year_counts
    }

@app.get("/dashboard")
def show_dashboard(request: Request):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse("/select-menu")  # 또는 "/login"
    return templates.TemplateResponse("dashboard.html", {"request": request, "user": user})

@app.get("/api/latest-updates")
def get_latest_updated_issues(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    username = user["username"]
    is_admin = user["role"] == "admin"

    subquery = (
        db.query(
            IssueDetail.issue_id,
            func.max(IssueDetail.date).label("latest_date")
        )
        .group_by(IssueDetail.issue_id)
        .subquery()
    )

    query = (
        db.query(Issue.id, Issue.issue_name, subquery.c.latest_date)
        .join(subquery, Issue.id == subquery.c.issue_id)
    )

    # ✅ 권한 처리 로직 동일하게 맞추기
    if not is_admin:
        query = query.filter(
            or_(
                Issue.is_hidden == False,
                and_(
                    Issue.is_hidden == True,
                    Issue.authorized_users != None,
                    Issue.authorized_users != "",
                    Issue.authorized_users.like(f"%,{username},%")
                )
            )
        )

    results = query.order_by(desc(subquery.c.latest_date)).limit(7).all()

    return [
        {
            "id": issue.id,
            "name": issue.issue_name,
            "updated_at": issue.latest_date.strftime("%Y-%m-%d")
        }
        for issue in results
    ]

router = APIRouter()

@router.get("/api/latest-updates")
def get_latest_updated_issues(db: Session = Depends(get_db)):
    subquery = (
        db.query(
            IssueDetail.issue_id,
            func.max(IssueDetail.date).label("latest_date")
        )
        .group_by(IssueDetail.issue_id)
        .subquery()
    )

    results = (
        db.query(Issue.id, Issue.issue_name, subquery.c.latest_date)
        .join(subquery, Issue.id == subquery.c.issue_id)
        .order_by(desc(subquery.c.latest_date))
        .limit(7)
        .all()
    )

    return [
        {
            "id": issue.id,
            "name": issue.issue_name,
            "updated_at": issue.latest_date.strftime("%Y-%m-%d")
        } for issue in results
    ]

@app.get("/api/schedules")
def get_schedules(db: Session = Depends(get_db)):
    today = datetime.today().date()
    schedules = (
        db.query(Schedule)
        .filter(or_(Schedule.end_date == None, Schedule.end_date >= today))  # ✅ 오늘 이후만
        .order_by(Schedule.start_date)
        .all()
    )
    result = []
    for s in schedules:
        result.append({
            "id": s.id,  # ✅ id 추가
            "start_date": s.start_date.strftime("%Y-%m-%d"),
            "end_date": s.end_date.strftime("%Y-%m-%d") if s.end_date else None,
            "title": s.title,
            "location": s.location
        })
    return result

@app.post("/api/schedules/add")
def add_schedule(schedule: ScheduleCreate, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    new_schedule = Schedule(
        start_date=schedule.start_date,
        end_date=schedule.end_date,
        title=schedule.title,
        location=schedule.location
    )
    db.add(new_schedule)
    db.commit()
    db.refresh(new_schedule)
    return {"message": "일정이 추가되었습니다."}

@app.get("/schedule/add")
def schedule_add_page(request: Request, user: dict = Depends(is_admin)):
    return templates.TemplateResponse("add_schedule.html", {"request": request})

@app.post("/schedule/add")
def schedule_add(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(None),
    title: str = Form(...),
    location: str = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(is_admin)
):
    # 문자열 → date 객체로 변환
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    schedule = Schedule(
        start_date=start_date_obj,
        end_date=end_date_obj,
        title=title,
        location=location,
        created_at=datetime.utcnow()
    )
    db.add(schedule)
    db.commit()
    return RedirectResponse("/select-menu", status_code=303)

@app.post("/api/schedule/cleanup")
def cleanup_past_schedules(db: Session = Depends(get_db)):
    today = date.today()

    schedules = db.query(Schedule).filter(
        or_(
            and_(Schedule.end_date != None, Schedule.end_date < today),
            and_(Schedule.end_date == None, Schedule.start_date < today)
        )
    ).all()

    for s in schedules:
        # 🔁 삭제 전에 백업
        history = ScheduleHistory(
            start_date=s.start_date,
            end_date=s.end_date,
            title=s.title,
            location=s.location
        )
        db.add(history)
        db.delete(s)  # 기존 방식 그대로 삭제

    db.commit()
    return {"deleted": len(schedules)}

@app.post("/schedule/delete/{schedule_id}")
async def delete_schedule(schedule_id: int, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")

    # 🔁 삭제 전에 백업
    history = ScheduleHistory(
        start_date=schedule.start_date,
        end_date=schedule.end_date,
        title=schedule.title,
        location=schedule.location
    )
    db.add(history)

    db.delete(schedule)
    db.commit()
    return {"message": "일정이 삭제되었습니다."}

@app.get("/schedule/history", response_class=HTMLResponse)
def schedule_history_page(request: Request, db: Session = Depends(get_db), user: dict = Depends(is_admin)):
    history = db.query(ScheduleHistory).order_by(ScheduleHistory.deleted_at.desc()).all()
    return templates.TemplateResponse("schedule_history.html", {
        "request": request,
        "schedules": history
    })

@app.get("/subscribe-alert", response_class=HTMLResponse)
def show_subscribe_page(request: Request):
    return templates.TemplateResponse("subscribe_alert.html", {"request": request})


@app.post("/subscribe-alert")
def subscribe_alert(request: Request, email: str = Form(...), db: Session = Depends(get_db)):
    existing = db.query(IssueAlertSubscription).filter_by(email=email).first()
    if existing:
        return templates.TemplateResponse("subscribe_alert.html", {
            "request": request,
            "message": "⚠️ 이미 신청된 이메일입니다."
        })

    db.add(IssueAlertSubscription(email=email))
    db.commit()
    return templates.TemplateResponse("subscribe_alert.html", {
        "request": request,
        "message": "✅ 알림 신청이 완료되었습니다."
    })

@app.on_event("startup")
def start_scheduler():
    scheduler = BackgroundScheduler(timezone=pytz.timezone("Asia/Seoul"))
    scheduler.add_job(
        lambda: send_daily_update_emails(next(get_db())),
        CronTrigger(hour=16, minute=0)
    )
    scheduler.start()

@app.get("/test-send-alerts")
def test_send_alerts(db: Session = Depends(get_db)):
    send_daily_update_emails(db)
    return {"status": "테스트 메일 발송 완료"}

@app.get("/alert-subscribers", response_class=HTMLResponse)
def alert_subscribers(request: Request, db: Session = Depends(get_db)):
    # 관리자 체크
    user = request.session.get("user")
    if not user or user["role"] != "admin":
        return RedirectResponse("/login", status_code=302)

    subscribers = db.query(IssueAlertSubscription).all()
    return templates.TemplateResponse("alert_subscribers.html", {
        "request": request,
        "subscribers": subscribers
    })


@app.post("/alert-subscribers/delete/{subscriber_id}")
def delete_subscriber(subscriber_id: int, db: Session = Depends(get_db)):
    subscriber = db.query(IssueAlertSubscription).filter_by(id=subscriber_id).first()
    if subscriber:
        db.delete(subscriber)
        db.commit()
    return RedirectResponse("/alert-subscribers", status_code=303)